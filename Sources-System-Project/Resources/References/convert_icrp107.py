import os
import json
import openpyxl

excel_path = r'd:\Sources-System\Sources-System-Project\Resources\References\ICRP107_AnnexA_Radionuclide_Database_REVIEWED.xlsx'
output_json_path = r'd:\Sources-System\Sources-System-Project\Resources\References\icrp107_decay_index.json'

print(f"Loading Excel file: {excel_path}...")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['Radionuclides']

headers = [cell.value for cell in next(ws.iter_rows())]
print(f"Headers found: {headers}")

field_map = {
    'nuclide': 'Nuclide',
    'is_radioactive': 'IsRadioactive',
    'half_life_raw': 'HalfLifeRaw',
    'half_life_seconds': 'HalfLifeSeconds',
    'decay_mode_raw': 'DecayModeRaw',
    'alpha_energy_mev': 'AlphaEnergy_MeV',
    'electron_energy_mev': 'ElectronEnergy_MeV',
    'photon_energy_mev': 'PhotonEnergy_MeV',
    'total_energy_mev': 'TotalEnergy_MeV'
}

col_indices = {key: headers.index(col_name) for key, col_name in field_map.items()}

entries = []
stable_count = 0

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    nuclide_val = row[col_indices['nuclide']].value
    if not nuclide_val:
        continue
    
    is_radioactive_raw = row[col_indices['is_radioactive']].value
    # Exclude non-radioactive / stable nuclides (IsRadioactive == False)
    if is_radioactive_raw is False or str(is_radioactive_raw).strip().lower() == 'false':
        stable_count += 1
        continue

    nuclide_str = str(nuclide_val).strip()

    half_life_raw = row[col_indices['half_life_raw']].value
    half_life_raw_str = str(half_life_raw).strip() if half_life_raw is not None else ""

    half_life_sec = row[col_indices['half_life_seconds']].value
    try:
        half_life_sec = float(half_life_sec) if half_life_sec is not None else None
    except (ValueError, TypeError):
        half_life_sec = None

    decay_mode_raw = row[col_indices['decay_mode_raw']].value
    decay_mode_str = str(decay_mode_raw).strip() if decay_mode_raw is not None else ""

    def parse_float(val):
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    alpha_energy = parse_float(row[col_indices['alpha_energy_mev']].value)
    electron_energy = parse_float(row[col_indices['electron_energy_mev']].value)
    photon_energy = parse_float(row[col_indices['photon_energy_mev']].value)
    total_energy = parse_float(row[col_indices['total_energy_mev']].value)

    entry = {
        "nuclide": nuclide_str,
        "is_radioactive": True,
        "half_life_raw": half_life_raw_str,
        "half_life_seconds": half_life_sec,
        "decay_mode_raw": decay_mode_str,
        "alpha_energy_mev": alpha_energy,
        "electron_energy_mev": electron_energy,
        "photon_energy_mev": photon_energy,
        "total_energy_mev": total_energy,
        "source_reference": "ICRP Publication 107 (Annex A, 2008)"
    }
    entries.append(entry)

print(f"Filtered: {len(entries)} radioactive radionuclides kept, {stable_count} stable nuclides excluded.")
with open(output_json_path, 'w', encoding='utf-8') as f:
    json.dump(entries, f, ensure_ascii=False, indent=2)

print(f"Successfully saved to: {output_json_path}")
